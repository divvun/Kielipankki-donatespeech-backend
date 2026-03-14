#!/usr/bin/env python3
"""Convert an authoring workbook into runtime schedule/theme JSON files.

Expected workbook structure:

- `ScheduleMeta` sheet (required):
    - `scheduleId` (required)
    - Optional start/finish columns:
        - `start_title_fi`, `start_title_nb`, `start_body1_fi`, `start_body1_nb`,
            `start_body2_fi`, `start_body2_nb`, `start_imageUrl`
        - `finish_title_fi`, `finish_title_nb`, `finish_body1_fi`, `finish_body1_nb`,
            `finish_body2_fi`, `finish_body2_nb`, `finish_imageUrl`

- `Items` sheet (required): one row per schedule item.

- `ItemOptions` sheet (optional): option rows for prompt item types that need options.

- `Theme` sheet (optional): one theme row.

- `ThemeSchedules` sheet (optional): list of scheduleIds for the theme.

The converter writes generated JSON files to:
`content/<env>/schedules/<scheduleId>.json` and
`content/<env>/themes/<themeId>.json`.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import ValidationError

from models import (
    AudioMediaItem,
    ChoicePromptItem,
    FakeYleAudioMediaItem,
    FakeYleVideoMediaItem,
    ImageMediaItem,
    MultiChoicePromptItem,
    Schedule,
    SuperChoicePromptItem,
    TextContentItem,
    TextInputItem,
    TextMediaItem,
    Theme,
    VideoMediaItem,
    YleAudioMediaItem,
    YleVideoMediaItem,
)


class WorkbookStructureError(Exception):
    """Raised when workbook structure is incompatible with converter schema."""


class RowValidationError(Exception):
    """Raised when an individual row cannot be converted."""


ITEM_MODEL_MAP: dict[tuple[str, str], type] = {
    ("media", "audio"): AudioMediaItem,
    ("media", "video"): VideoMediaItem,
    ("media", "yle-audio"): YleAudioMediaItem,
    ("media", "yle-video"): YleVideoMediaItem,
    ("media", "fake-yle-audio"): FakeYleAudioMediaItem,
    ("media", "fake-yle-video"): FakeYleVideoMediaItem,
    ("media", "text-content"): TextContentItem,
    ("media", "image"): ImageMediaItem,
    ("media", "text"): TextMediaItem,
    ("prompt", "choice"): ChoicePromptItem,
    ("prompt", "multi-choice"): MultiChoicePromptItem,
    ("prompt", "super-choice"): SuperChoicePromptItem,
    ("prompt", "text"): TextInputItem,
}

PROMPT_TYPES_WITH_OPTIONS = {"choice", "multi-choice", "super-choice"}
PROMPT_TYPES_WITH_OTHER_ENTRY = {"multi-choice", "super-choice"}

REQUIRED_ITEM_COLUMNS = (
    "itemId",
    "kind",
    "itemType",
    "url",
    "isRecording",
    "default_title_fi",
    "default_title_nb",
    "default_body1_fi",
    "default_body1_nb",
    "default_body2_fi",
    "default_body2_nb",
)


@dataclass
class SkippedRow:
    sheet: str
    row: int
    reason: str


@dataclass
class ConversionResult:
    workbook: str
    schedule_id: str
    schedule_path: str
    theme_id: str | None
    theme_path: str | None
    skipped_rows: list[SkippedRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "workbook": self.workbook,
            "schedule_id": self.schedule_id,
            "schedule_path": self.schedule_path,
            "theme_id": self.theme_id,
            "theme_path": self.theme_path,
            "skipped_rows": [asdict(item) for item in self.skipped_rows],
            "warnings": self.warnings,
        }


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    text = str(value).strip()
    return text if text else None


def _text_or_empty(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        if value in (0, 1):
            return bool(value)
        raise ValueError("boolean value must be 0 or 1 when numeric")

    text = _normalize_text(value)
    if text is None:
        raise ValueError("boolean value is missing")

    lowered = text.lower()
    if lowered in {"true", "1", "yes", "y"}:
        return True
    if lowered in {"false", "0", "no", "n"}:
        return False

    raise ValueError(f"invalid boolean value: {text}")


def _parse_int(value: Any, field_name: str) -> int:
    if value is None:
        raise ValueError(f"{field_name} is missing")

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"{field_name} must be an integer")

    text = _normalize_text(value)
    if text is None:
        raise ValueError(f"{field_name} is missing")

    try:
        return int(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be an integer") from exc


def _format_validation_error(exc: ValidationError) -> str:
    details: list[str] = []
    for err in exc.errors():
        location = ".".join(str(part) for part in err.get("loc", ()))
        message = err.get("msg", "validation error")
        details.append(f"{location}: {message}" if location else message)
    return "; ".join(details[:3])


def _read_sheet_rows(
    workbook: Any,
    sheet_name: str,
    required_columns: tuple[str, ...],
    optional: bool = False,
) -> list[tuple[int, dict[str, Any]]]:
    if sheet_name not in workbook.sheetnames:
        if optional:
            return []
        raise WorkbookStructureError(f"Missing required sheet: {sheet_name}")

    sheet = workbook[sheet_name]
    raw_headers = [cell.value for cell in sheet[1]]
    header_columns: list[tuple[int, str]] = []
    seen_headers: dict[str, int] = {}

    for index, value in enumerate(raw_headers, start=1):
        header = _normalize_text(value)
        if header is None:
            continue
        if header in seen_headers:
            raise WorkbookStructureError(
                f"Duplicate column '{header}' in sheet {sheet_name}"
            )
        seen_headers[header] = index
        header_columns.append((index, header))

    if not header_columns:
        raise WorkbookStructureError(f"Sheet {sheet_name} has no header row")

    missing = [column for column in required_columns if column not in seen_headers]
    if missing:
        missing_str = ", ".join(missing)
        raise WorkbookStructureError(
            f"Sheet {sheet_name} is missing required columns: {missing_str}"
        )

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number in range(2, sheet.max_row + 1):
        row: dict[str, Any] = {}
        has_data = False
        for col_idx, header in header_columns:
            value = sheet.cell(row=row_number, column=col_idx).value
            row[header] = value
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            has_data = True
        if has_data:
            rows.append((row_number, row))

    return rows


def _row_has_data(row: dict[str, Any], keys: tuple[str, ...]) -> bool:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def _localized_optional(row: dict[str, Any], prefix: str) -> dict[str, str] | None:
    fi = _normalize_text(row.get(f"{prefix}_fi"))
    nb = _normalize_text(row.get(f"{prefix}_nb"))

    if fi is None and nb is None:
        return None
    if fi is None or nb is None:
        raise RowValidationError(
            f"Both {prefix}_fi and {prefix}_nb are required when one is provided"
        )

    return {"fi": fi, "nb": nb}


def _localized_required(row: dict[str, Any], prefix: str) -> dict[str, str]:
    return {
        "fi": _text_or_empty(row.get(f"{prefix}_fi")),
        "nb": _text_or_empty(row.get(f"{prefix}_nb")),
    }


def _media_state_from_row(
    row: dict[str, Any],
    prefix: str,
    required: bool,
) -> dict[str, Any] | None:
    keys = (
        f"{prefix}_title_fi",
        f"{prefix}_title_nb",
        f"{prefix}_body1_fi",
        f"{prefix}_body1_nb",
        f"{prefix}_body2_fi",
        f"{prefix}_body2_nb",
        f"{prefix}_imageUrl",
    )
    state_present = _row_has_data(row, keys)
    if not required and not state_present:
        return None

    return {
        "title": _localized_required(row, f"{prefix}_title"),
        "body1": _localized_required(row, f"{prefix}_body1"),
        "body2": _localized_required(row, f"{prefix}_body2"),
        "imageUrl": _normalize_text(row.get(f"{prefix}_imageUrl")),
    }


def _collect_item_options(
    option_rows: list[tuple[int, dict[str, Any]]],
    skipped_rows: list[SkippedRow],
    warnings: list[str],
) -> dict[str, list[dict[str, str]]]:
    options_by_item: dict[str, list[tuple[int, dict[str, str]]]] = {}
    seen_indices: dict[str, set[int]] = {}

    for row_number, row in option_rows:
        try:
            item_id = _normalize_text(row.get("itemId"))
            if item_id is None:
                raise RowValidationError("itemId is required")

            option_index = _parse_int(row.get("optionIndex"), "optionIndex")
            if option_index < 0:
                raise RowValidationError("optionIndex must be >= 0")

            fi = _normalize_text(row.get("option_fi"))
            nb = _normalize_text(row.get("option_nb"))
            if fi is None and nb is None:
                raise RowValidationError("option_fi and option_nb are required")
            if fi is None or nb is None:
                raise RowValidationError(
                    "Both option_fi and option_nb are required when one is provided"
                )

            options_by_item.setdefault(item_id, [])
            seen_indices.setdefault(item_id, set())

            if option_index in seen_indices[item_id]:
                warnings.append(
                    f"Duplicate optionIndex {option_index} for item {item_id}; "
                    f"ItemOptions row {row_number} was ignored"
                )
                continue

            seen_indices[item_id].add(option_index)
            options_by_item[item_id].append((option_index, {"fi": fi, "nb": nb}))

        except (RowValidationError, ValueError) as exc:
            skipped_rows.append(
                SkippedRow(sheet="ItemOptions", row=row_number, reason=str(exc))
            )

    normalized: dict[str, list[dict[str, str]]] = {}
    for item_id, indexed_options in options_by_item.items():
        normalized[item_id] = [
            option for _, option in sorted(indexed_options, key=lambda pair: pair[0])
        ]
    return normalized


def _resolve_item_order(row: dict[str, Any], row_number: int) -> int:
    order_raw = _normalize_text(row.get("order"))
    if order_raw is None:
        return row_number

    order = _parse_int(order_raw, "order")
    if order < 0:
        raise RowValidationError("order must be >= 0")
    return order


def _build_item_from_row(
    row: dict[str, Any],
    row_number: int,
    options_by_item: dict[str, list[dict[str, str]]],
    warnings: list[str],
) -> tuple[int, dict[str, Any]]:
    kind = _normalize_text(row.get("kind"))
    item_type = _normalize_text(row.get("itemType"))
    item_id = _normalize_text(row.get("itemId"))
    url = _normalize_text(row.get("url"))

    if kind is None:
        raise RowValidationError("kind is required")
    if item_type is None:
        raise RowValidationError("itemType is required")
    if item_id is None:
        raise RowValidationError("itemId is required")
    if url is None:
        raise RowValidationError("url is required")

    kind = kind.lower()
    item_type = item_type.lower()

    if kind == "prompt" and item_type == "text-input":
        warnings.append(
            f"Items row {row_number}: prompt/text-input normalized to prompt/text"
        )
        item_type = "text"

    key = (kind, item_type)
    model_cls = ITEM_MODEL_MAP.get(key)
    if model_cls is None:
        valid_types = ", ".join(
            sorted(f"{known_kind}/{known_type}" for known_kind, known_type in ITEM_MODEL_MAP)
        )
        raise RowValidationError(
            f"Unsupported kind/itemType combination '{kind}/{item_type}'. "
            f"Valid values: {valid_types}"
        )

    try:
        is_recording = _parse_bool(row.get("isRecording"))
    except ValueError as exc:
        raise RowValidationError(str(exc)) from exc

    item_data: dict[str, Any] = {
        "kind": kind,
        "itemType": item_type,
        "itemId": item_id,
        "url": url,
        "isRecording": is_recording,
        "default": _media_state_from_row(row, "default", required=True),
    }

    type_id = _normalize_text(row.get("typeId"))
    if type_id is not None:
        item_data["typeId"] = type_id

    if kind == "media":
        item_data["options"] = []
        for prefix in ("start", "recording", "finish"):
            state = _media_state_from_row(row, prefix, required=False)
            if state is not None:
                item_data[prefix] = state

        meta_title = _localized_optional(row, "metaTitle")
        if meta_title is not None:
            item_data["metaTitle"] = meta_title

    if kind == "prompt":
        if item_type in PROMPT_TYPES_WITH_OPTIONS:
            options = options_by_item.get(item_id, [])
            if not options:
                raise RowValidationError(
                    f"Prompt item {item_id} requires at least one option in ItemOptions"
                )
            item_data["options"] = options
        else:
            item_data["options"] = []

        if item_type == "text":
            meta_title = _localized_optional(row, "metaTitle")
            if meta_title is not None:
                item_data["metaTitle"] = meta_title

        if item_type == "multi-choice":
            other_answer = _localized_optional(row, "otherAnswer")
            if other_answer is not None:
                item_data["otherAnswer"] = other_answer

        if item_type in PROMPT_TYPES_WITH_OTHER_ENTRY:
            other_entry = _localized_optional(row, "otherEntryLabel")
            if other_entry is not None:
                item_data["otherEntryLabel"] = other_entry

    try:
        parsed = model_cls(**item_data)
    except ValidationError as exc:
        raise RowValidationError(_format_validation_error(exc)) from exc

    order = _resolve_item_order(row, row_number)
    return order, parsed.model_dump(exclude_none=True)


def _build_schedule(
    workbook: Any,
    items: list[dict[str, Any]],
    warnings: list[str],
) -> tuple[str, Schedule]:
    schedule_rows = _read_sheet_rows(
        workbook=workbook,
        sheet_name="ScheduleMeta",
        required_columns=("scheduleId",),
        optional=False,
    )

    if not schedule_rows:
        raise WorkbookStructureError("ScheduleMeta must contain at least one data row")

    if len(schedule_rows) > 1:
        warnings.append("ScheduleMeta has multiple rows; only the first row was used")

    row_number, row = schedule_rows[0]
    schedule_id = _normalize_text(row.get("scheduleId"))
    if schedule_id is None:
        raise WorkbookStructureError(
            f"ScheduleMeta row {row_number}: scheduleId is required"
        )

    schedule_data: dict[str, Any] = {
        "scheduleId": schedule_id,
        "items": items,
    }

    for prefix in ("start", "finish"):
        state = _media_state_from_row(row, prefix, required=False)
        if state is not None:
            schedule_data[prefix] = state

    try:
        schedule_model = Schedule(**schedule_data)
    except ValidationError as exc:
        raise WorkbookStructureError(
            f"ScheduleMeta row {row_number} failed validation: "
            f"{_format_validation_error(exc)}"
        ) from exc

    return schedule_id, schedule_model


def _build_theme(
    workbook: Any,
    fallback_schedule_id: str,
    skipped_rows: list[SkippedRow],
    warnings: list[str],
) -> tuple[str | None, Theme | None]:
    theme_rows = _read_sheet_rows(
        workbook=workbook,
        sheet_name="Theme",
        required_columns=(
            "themeId",
            "title_fi",
            "title_nb",
            "body1_fi",
            "body1_nb",
            "body2_fi",
            "body2_nb",
        ),
        optional=True,
    )

    if not theme_rows:
        return None, None

    if len(theme_rows) > 1:
        warnings.append("Theme has multiple rows; only the first row was used")

    row_number, row = theme_rows[0]
    try:
        theme_id = _normalize_text(row.get("themeId"))
        if theme_id is None:
            raise RowValidationError("themeId is required")

        schedule_rows = _read_sheet_rows(
            workbook=workbook,
            sheet_name="ThemeSchedules",
            required_columns=("scheduleId",),
            optional=True,
        )

        schedule_ids: list[str] = []
        for schedule_row_number, schedule_row in schedule_rows:
            schedule_id = _normalize_text(schedule_row.get("scheduleId"))
            if schedule_id is None:
                skipped_rows.append(
                    SkippedRow(
                        sheet="ThemeSchedules",
                        row=schedule_row_number,
                        reason="scheduleId is required",
                    )
                )
                continue
            if schedule_id not in schedule_ids:
                schedule_ids.append(schedule_id)

        if not schedule_ids:
            schedule_ids = [fallback_schedule_id]

        theme_data: dict[str, Any] = {
            "title": _localized_required(row, "title"),
            "body1": _localized_required(row, "body1"),
            "body2": _localized_required(row, "body2"),
            "scheduleIds": schedule_ids,
        }

        image = _normalize_text(row.get("image"))
        if image is not None:
            theme_data["image"] = image

        theme_model = Theme(**theme_data)
        return theme_id, theme_model

    except (RowValidationError, ValidationError) as exc:
        reason = str(exc)
        if isinstance(exc, ValidationError):
            reason = _format_validation_error(exc)
        skipped_rows.append(SkippedRow(sheet="Theme", row=row_number, reason=reason))
        return None, None


def convert_workbook(
    workbook_path: Path,
    output_env: str = "dev",
    content_root: Path = Path("content"),
    strict: bool = False,
) -> ConversionResult:
    workbook = load_workbook(filename=workbook_path, data_only=True)
    skipped_rows: list[SkippedRow] = []
    warnings: list[str] = []

    option_rows = _read_sheet_rows(
        workbook=workbook,
        sheet_name="ItemOptions",
        required_columns=("itemId", "optionIndex", "option_fi", "option_nb"),
        optional=True,
    )
    options_by_item = _collect_item_options(
        option_rows=option_rows,
        skipped_rows=skipped_rows,
        warnings=warnings,
    )

    item_rows = _read_sheet_rows(
        workbook=workbook,
        sheet_name="Items",
        required_columns=REQUIRED_ITEM_COLUMNS,
        optional=False,
    )

    parsed_items: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in item_rows:
        try:
            parsed_items.append(
                _build_item_from_row(
                    row=row,
                    row_number=row_number,
                    options_by_item=options_by_item,
                    warnings=warnings,
                )
            )
        except RowValidationError as exc:
            skipped_rows.append(SkippedRow(sheet="Items", row=row_number, reason=str(exc)))

    if not parsed_items:
        raise WorkbookStructureError("No valid Items rows could be converted")

    sorted_items = [item for _, item in sorted(parsed_items, key=lambda pair: pair[0])]
    schedule_id, schedule_model = _build_schedule(
        workbook=workbook,
        items=sorted_items,
        warnings=warnings,
    )

    theme_id, theme_model = _build_theme(
        workbook=workbook,
        fallback_schedule_id=schedule_id,
        skipped_rows=skipped_rows,
        warnings=warnings,
    )

    if strict and skipped_rows:
        raise WorkbookStructureError(
            f"Strict mode enabled and {len(skipped_rows)} row(s) were skipped"
        )

    schedule_dir = content_root / output_env / "schedules"
    schedule_dir.mkdir(parents=True, exist_ok=True)
    schedule_path = schedule_dir / f"{schedule_id}.json"

    with open(schedule_path, "w", encoding="utf-8") as schedule_file:
        json.dump(schedule_model.model_dump(exclude_none=True), schedule_file, indent=2, ensure_ascii=False)
        schedule_file.write("\n")

    theme_path: Path | None = None
    if theme_id is not None and theme_model is not None:
        theme_dir = content_root / output_env / "themes"
        theme_dir.mkdir(parents=True, exist_ok=True)
        theme_path = theme_dir / f"{theme_id}.json"
        with open(theme_path, "w", encoding="utf-8") as theme_file:
            json.dump(theme_model.model_dump(exclude_none=True), theme_file, indent=2, ensure_ascii=False)
            theme_file.write("\n")

    return ConversionResult(
        workbook=str(workbook_path),
        schedule_id=schedule_id,
        schedule_path=str(schedule_path),
        theme_id=theme_id,
        theme_path=str(theme_path) if theme_path else None,
        skipped_rows=skipped_rows,
        warnings=warnings,
    )


def _write_report(report_path: Path, result: ConversionResult) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as report_file:
        json.dump(result.as_dict(), report_file, indent=2, ensure_ascii=False)
        report_file.write("\n")


def _build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert a content authoring workbook into schedule/theme JSON files"
    )
    parser.add_argument("workbook", type=Path, help="Path to source workbook (.xlsx)")
    parser.add_argument(
        "--output-env",
        choices=("dev", "prod"),
        default="dev",
        help="Target content environment directory (default: dev)",
    )
    parser.add_argument(
        "--content-root",
        type=Path,
        default=Path("content"),
        help="Root content directory (default: ./content)",
    )
    parser.add_argument(
        "--report",
        type=Path,
        default=None,
        help="Optional JSON report output path",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail when any row is skipped",
    )
    return parser


def main() -> int:
    parser = _build_argument_parser()
    args = parser.parse_args()

    try:
        result = convert_workbook(
            workbook_path=args.workbook,
            output_env=args.output_env,
            content_root=args.content_root,
            strict=args.strict,
        )
    except WorkbookStructureError as exc:
        print(f"ERROR: {exc}")
        return 1
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}")
        return 1

    print(f"Generated schedule: {result.schedule_path}")
    if result.theme_path:
        print(f"Generated theme: {result.theme_path}")
    else:
        print("Generated theme: skipped (no valid Theme row)")

    if result.skipped_rows:
        print(f"Skipped rows: {len(result.skipped_rows)}")
        for skipped in result.skipped_rows:
            print(f"  - {skipped.sheet} row {skipped.row}: {skipped.reason}")

    if result.warnings:
        print(f"Warnings: {len(result.warnings)}")
        for warning in result.warnings:
            print(f"  - {warning}")

    if args.report:
        _write_report(args.report, result)
        print(f"Report: {args.report}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
