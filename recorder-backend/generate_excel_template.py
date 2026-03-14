#!/usr/bin/env python3
"""Generate a starter Excel workbook for schedule/theme authoring.

The generated workbook is compatible with `convert_excel_to_json.py` and includes
all required sheets and column headers.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]


SHEET_SCHEDULE_META = "ScheduleMeta"
SHEET_ITEMS = "Items"
SHEET_ITEM_OPTIONS = "ItemOptions"
SHEET_THEME = "Theme"
SHEET_THEME_SCHEDULES = "ThemeSchedules"


SCHEDULE_META_HEADERS = [
    "scheduleId",
    "start_title_fi",
    "start_title_nb",
    "start_body1_fi",
    "start_body1_nb",
    "start_body2_fi",
    "start_body2_nb",
    "start_url",
    "finish_title_fi",
    "finish_title_nb",
    "finish_body1_fi",
    "finish_body1_nb",
    "finish_body2_fi",
    "finish_body2_nb",
    "finish_url",
]


ITEM_HEADERS = [
    "order",
    "itemId",
    "kind",
    "itemType",
    "url",
    "typeId",
    "isRecording",
    "start_title_fi",
    "start_title_nb",
    "start_body1_fi",
    "start_body1_nb",
    "start_body2_fi",
    "start_body2_nb",
    "start_url",
    "recording_title_fi",
    "recording_title_nb",
    "recording_body1_fi",
    "recording_body1_nb",
    "recording_body2_fi",
    "recording_body2_nb",
    "recording_url",
    "finish_title_fi",
    "finish_title_nb",
    "finish_body1_fi",
    "finish_body1_nb",
    "finish_body2_fi",
    "finish_body2_nb",
    "finish_url",
    "otherAnswer_fi",
    "otherAnswer_nb",
    "otherEntryLabel_fi",
    "otherEntryLabel_nb",
]


ITEM_OPTIONS_HEADERS = [
    "itemId",
    "optionIndex",
    "option_fi",
    "option_nb",
]


THEME_HEADERS = [
    "themeId",
    "title_fi",
    "title_nb",
    "body1_fi",
    "body1_nb",
    "body2_fi",
    "body2_nb",
    "image",
]


THEME_SCHEDULE_HEADERS = [
    "scheduleId",
]


ITEM_KINDS = ["media", "prompt"]
ITEM_TYPES = [
    "audio",
    "video",
    "yle-audio",
    "yle-video",
    "fake-yle-audio",
    "fake-yle-video",
    "text-content",
    "image",
    "text",
    "choice",
    "multi-choice",
    "super-choice",
]


def _style_header_row(sheet) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def _set_column_widths(sheet, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        width = max(14, min(42, len(header) + 4))
        sheet.column_dimensions[chr(64 + idx) if idx <= 26 else sheet.cell(row=1, column=idx).column_letter].width = width


def _prepare_sheet(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    _style_header_row(sheet)
    _set_column_widths(sheet, headers)
    sheet.freeze_panes = "A2"


def _add_item_sheet_validations(items_sheet) -> None:
    kind_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(ITEM_KINDS) + '"',
        allow_blank=False,
    )
    items_sheet.add_data_validation(kind_validation)
    kind_validation.add("C2:C2000")

    item_type_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(ITEM_TYPES) + '"',
        allow_blank=False,
    )
    items_sheet.add_data_validation(item_type_validation)
    item_type_validation.add("D2:D2000")

    bool_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
    )
    items_sheet.add_data_validation(bool_validation)
    bool_validation.add("G2:G2000")


def create_template_workbook(output_path: Path, overwrite: bool = False) -> Path:
    """Create a starter workbook with all required sheets and headers."""
    if output_path.exists() and not overwrite:
        raise FileExistsError(
            f"Output file already exists: {output_path}. Use --force to overwrite."
        )

    workbook = Workbook()

    schedule_meta = workbook.active
    schedule_meta.title = SHEET_SCHEDULE_META
    _prepare_sheet(schedule_meta, SCHEDULE_META_HEADERS)
    schedule_meta.append(["replace-schedule-id"] + [""] * (len(SCHEDULE_META_HEADERS) - 1))

    items = workbook.create_sheet(SHEET_ITEMS)
    _prepare_sheet(items, ITEM_HEADERS)
    _add_item_sheet_validations(items)
    items.append(
        [
            1,
            "replace-item-id-media-001",
            "media",
            "image",
            "https://example.org/image.jpg",
            "image/jpeg",
            "TRUE",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    items.append(
        [
            2,
            "replace-item-id-choice-001",
            "prompt",
            "choice",
            "https://example.org/prompt-image.jpg",
            "",
            "FALSE",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    item_options = workbook.create_sheet(SHEET_ITEM_OPTIONS)
    _prepare_sheet(item_options, ITEM_OPTIONS_HEADERS)
    item_options.append(["replace-item-id-choice-001", 0, "", ""])
    item_options.append(["replace-item-id-choice-001", 1, "", ""])

    theme = workbook.create_sheet(SHEET_THEME)
    _prepare_sheet(theme, THEME_HEADERS)
    theme.append(["replace-theme-id", "", "", "", "", "", "", ""])

    theme_schedules = workbook.create_sheet(SHEET_THEME_SCHEDULES)
    _prepare_sheet(theme_schedules, THEME_SCHEDULE_HEADERS)
    theme_schedules.append(["replace-schedule-id"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate a starter workbook for recorder content authoring"
    )
    parser.add_argument(
        "output",
        nargs="?",
        type=Path,
        default=Path("content-authoring-template.xlsx"),
        help="Output workbook path (default: content-authoring-template.xlsx)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite output file if it already exists",
    )
    return parser


def main() -> int:
    parser = _build_argument_parser()
    args = parser.parse_args()

    try:
        path = create_template_workbook(output_path=args.output, overwrite=args.force)
    except FileExistsError as exc:
        print(f"ERROR: {exc}")
        return 1

    print(f"Created workbook template: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
