#!/usr/bin/env python3
"""Convert theme worksheets from XLSX to language-specific theme JSON files.

Each worksheet is treated as one theme and exported to:
  ../recorder-content/dev/themes/<theme-id>/<lang>.json

The worksheet hierarchy follows the layout used in example-theme.csv and
export_theme_json_to_xlsx.py.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

LANGUAGE_EXPORT_MAP = {
    "north": "se",
    "south": "sma",
    "lule": "smj",
    "inari": "smn",
    "skolt": "sms", "ume": "sju", "pite": "sje",
}

IGNORED_LANGUAGES = {"bokmål", "suomi", "svenska"}

DRIVE_FILENAME_MAP = {
    "https://drive.google.com/file/d/1rRKcrEKdUXups7DF-MB1De42iSQTx6Ch/": "foto21_svt.jpg",
    "https://drive.google.com/file/d/1dHOykpGjJUz21uQ7JFBEk8tT3MRe6338/": "foto3_svt.jpg",
    "https://drive.google.com/file/d/1yTn-13QULyCPPpUdYWHnHI4TkGOwYKb4/": "foto_svt.jpg",
    "https://drive.google.com/file/d/1cfUoo9vByYjn2Epv3ZEOCLvtD_NSR63p/": "foto_michael_koneckiy.jpg",
    "https://drive.google.com/file/d/1b-Vp5Jd0S-z91gyqvZ6Q72hr3FQvOkV-": "foto26_svt.jpg",
    "https://drive.google.com/file/d/1D6XO_TR4dXHB2UjHM2cIEcDnUukv23t5/": "foto_tom_fisk.jpg",
    "https://drive.google.com/file/d/1abH6JOutlASs8Z2PzlipPbzaoSp6q7BR/": "foto_alina-bystrova.jpg",
    "https://drive.google.com/file/d/1FJ26SNDYPW8KoIwXrx88ZIPuH3APXpca/": "foto9_svt.jpg",
    "https://drive.google.com/file/d/1Uaebu4NumHUaCL5MESfkLQM_dyIi903R/": "foto_channnsy.jpg",
    "https://drive.google.com/file/d/1V1FeND4asOnNaSfkJnN3ksHmiO-bURG9/": "foto_djimmer-koster.jpg",
    "https://drive.google.com/file/d/17qWXXwLkl46--69Ksd02NL5--zJhLfi6/": "foto-anna-shvets.jpg",
    "https://drive.google.com/file/d/1sfpiD6oQJjvMtvAP6k9b2TAJxpxlbZ2G/": "foto2_svt.jpg",
    "https://drive.google.com/file/d/1uhM2LGlk4JYlO72ZuHH4yeIZPGHxa4a2/": "foto18_svt.jpg",
    "https://drive.google.com/file/d/1lYW7jkrrg5-s6I5OP6oj9OGUd2MRXJZG/": "illustrator_sunna_kuoljok_inga_wiktoria_pave.jpg",
    "https://drive.google.com/file/d/11yiBEMcUoDR9VhGenYA0t19q-lnAcCBs/": "foto14_svt.jpg",
    "https://drive.google.com/file/d/1GVeTpmKGbGrgEuJItRpeL-VgCjIRmHaP/": "foto13_svt.jpg",
    "https://drive.google.com/file/d/1RHd4g_LX7Au0BEN3OGPIYN4C823bpPY6/": "foto4_svt.jpg",
    "https://drive.google.com/file/d/1siRdn3n_jxgSzt6AL9N01Z69WxpqiT6B/": "madssuhrpettersen3.jpg",
    "https://drive.google.com/file/d/1Gh-voPSwncgNwPKv25ZpP8BMt_n2_v88/": "madssuhrpettersen.jpg",
    "https://drive.google.com/file/d/13SQh1cgAe2PDuNHDr-8HNPd_j6TEwXzl/": "foto15_svt.jpg",
    "https://drive.google.com/file/d/1whBnb_IlLoy7E011ljdCD5mJY-dPzJtm/": "jill-wellington.jpg",
    "https://drive.google.com/file/d/107hNkcJdaCwhGw8KEJ8UUaaCEjAD4UkA/": "foto10_svt.jpg",
    "https://drive.google.com/file/d/1qnP1Ax9Zf5vTz3ke6hlgmFpUgnmCAQ39/": "sami_beassas_markanat-mads_suhr_pettersen.jpg",
}


@dataclass
class ParsedState:
    url: str
    title: dict[str, str]
    body1: dict[str, str]
    body2: dict[str, str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert XLSX theme sheets to content/dev/themes/<theme-id>/<lang>.json"
    )
    parser.add_argument(
        "input_xlsx",
        type=Path,
        help="Input workbook path, for example uit-export.xlsx",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("../recorder-content/dev/themes"),
        help="Output themes root directory (default: ../recorder-content/dev/themes)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing output files",
    )
    return parser.parse_args()


def _normalize_label(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _row_values(ws: Any) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _find_row_by_label(rows: list[list[Any]], label: str, start_idx: int = 0) -> int:
    for idx in range(start_idx, len(rows)):
        row = rows[idx]
        for cell in row:
            if _normalize_label(cell) == label:
                return idx
    return -1


def _find_cell_value_in_row(row: list[Any], key: str) -> str:
    for idx, cell in enumerate(row):
        if _normalize_label(cell) == key:
            if idx + 1 < len(row):
                return _normalize_label(row[idx + 1])
            return ""
    return ""


def _parse_localized_block(
    rows: list[list[Any]],
    label_row_idx: int,
    label: str,
) -> tuple[dict[str, str], int]:
    label_row = rows[label_row_idx]
    if _normalize_label(label_row[0] if label_row else "") != label:
        # Also allow indented labels.
        first_idx = next(
            (i for i, cell in enumerate(label_row) if _normalize_label(cell)), -1
        )
        if first_idx == -1 or _normalize_label(label_row[first_idx]) != label:
            return {}, label_row_idx + 1
        label_col = first_idx
    else:
        label_col = 0

    header_idx = label_row_idx + 1
    value_idx = label_row_idx + 2
    if header_idx >= len(rows) or value_idx >= len(rows):
        return {}, min(len(rows), label_row_idx + 1)

    header_row = rows[header_idx]
    value_row = rows[value_idx]

    langs_by_col: dict[int, str] = {}
    for col_idx in range(label_col + 1, len(header_row)):
        lang = _normalize_label(header_row[col_idx]).lower()
        if lang:
            langs_by_col[col_idx] = lang

    localized: dict[str, str] = {}
    for col_idx, lang in langs_by_col.items():
        value = _normalize_label(value_row[col_idx] if col_idx < len(value_row) else "")
        localized[lang] = value

    return localized, value_idx + 1


def _parse_state(
    rows: list[list[Any]], state_label_row_idx: int
) -> tuple[ParsedState, int]:
    idx = state_label_row_idx + 1
    url = ""
    title: dict[str, str] = {}
    body1: dict[str, str] = {}
    body2: dict[str, str] = {}

    while idx < len(rows):
        row = rows[idx]
        first_idx = next(
            (i for i, cell in enumerate(row) if _normalize_label(cell)), -1
        )
        if first_idx == -1:
            idx += 1
            continue

        label = _normalize_label(row[first_idx])

        if first_idx == 0 and label in {
            "start",
            "finish",
            "items",
            "scheduleId",
            "themeId",
            "url",
        }:
            break
        if first_idx == 1 and label in {"itemId", "start", "recording", "finish"}:
            break

        if label == "url":
            url = _normalize_label(
                row[first_idx + 1] if first_idx + 1 < len(row) else ""
            )
            idx += 1
            continue
        if label == "title":
            title, idx = _parse_localized_block(rows, idx, "title")
            continue
        if label == "body1":
            body1, idx = _parse_localized_block(rows, idx, "body1")
            continue
        if label == "body2":
            body2, idx = _parse_localized_block(rows, idx, "body2")
            continue

        idx += 1

    return ParsedState(url=url, title=title, body1=body1, body2=body2), idx


def _parse_item(
    rows: list[list[Any]], item_id_row_idx: int
) -> tuple[dict[str, Any], int]:
    row = rows[item_id_row_idx]
    item_id = _find_cell_value_in_row(row, "itemId")
    item: dict[str, Any] = {
        "itemId": item_id,
        "kind": "media",
        "itemType": "image",
        "typeId": None,
    }

    idx = item_id_row_idx + 1
    while idx < len(rows):
        cur = rows[idx]
        first_idx = next(
            (i for i, cell in enumerate(cur) if _normalize_label(cell)), -1
        )
        if first_idx == -1:
            idx += 1
            continue

        label = _normalize_label(cur[first_idx])

        # Next item or top-level section.
        if first_idx == 1 and label == "itemId":
            break
        if first_idx == 0 and label in {
            "themeId",
            "scheduleId",
            "start",
            "finish",
            "items",
            "url",
            "title",
        }:
            break

        if first_idx == 1 and label in {"kind", "itemType", "typeId"}:
            value = _normalize_label(
                cur[first_idx + 1] if first_idx + 1 < len(cur) else ""
            )
            if label == "typeId":
                item["typeId"] = value or None
            else:
                item[label] = value or item[label]
            idx += 1
            continue

        if first_idx == 1 and label in {"start", "recording", "finish"}:
            parsed_state, next_idx = _parse_state(rows, idx)
            item[label] = parsed_state
            idx = next_idx
            continue

        idx += 1

    item["isRecording"] = bool(item.get("recording"))
    return item, idx


def _parse_sheet(ws: Any) -> tuple[str, str, dict[str, str], dict[str, Any]]:
    rows = _row_values(ws)

    theme_id_row = _find_row_by_label(rows, "themeId")
    if theme_id_row < 0:
        raise ValueError(f"Sheet '{ws.title}' does not contain a themeId row")

    theme_id = _find_cell_value_in_row(rows[theme_id_row], "themeId")
    if not theme_id:
        raise ValueError(f"Sheet '{ws.title}' has an empty themeId value")

    url_row = _find_row_by_label(rows, "url", start_idx=theme_id_row)
    theme_url = _find_cell_value_in_row(rows[url_row], "url") if url_row >= 0 else ""

    title_row = _find_row_by_label(rows, "title", start_idx=theme_id_row)
    theme_title, _ = (
        _parse_localized_block(rows, title_row, "title") if title_row >= 0 else ({}, 0)
    )

    schedule_id_row = _find_row_by_label(rows, "scheduleId", start_idx=theme_id_row)
    schedule_id = (
        _find_cell_value_in_row(rows[schedule_id_row], "scheduleId")
        if schedule_id_row >= 0
        else ""
    )

    start_row = _find_row_by_label(
        rows,
        "start",
        start_idx=schedule_id_row if schedule_id_row >= 0 else theme_id_row,
    )
    finish_row = _find_row_by_label(
        rows, "finish", start_idx=start_row + 1 if start_row >= 0 else theme_id_row
    )

    start_state, _ = (
        _parse_state(rows, start_row)
        if start_row >= 0
        else (ParsedState("", {}, {}, {}), 0)
    )
    finish_state, _ = (
        _parse_state(rows, finish_row)
        if finish_row >= 0
        else (ParsedState("", {}, {}, {}), 0)
    )

    items_row = _find_row_by_label(
        rows, "items", start_idx=finish_row if finish_row >= 0 else theme_id_row
    )
    items: list[dict[str, Any]] = []

    if items_row >= 0:
        idx = items_row + 1
        while idx < len(rows):
            row = rows[idx]
            first_idx = next(
                (i for i, cell in enumerate(row) if _normalize_label(cell)), -1
            )
            if first_idx == -1:
                idx += 1
                continue

            label = _normalize_label(row[first_idx])
            if first_idx == 1 and label == "itemId":
                parsed_item, next_idx = _parse_item(rows, idx)
                items.append(parsed_item)
                idx = next_idx
                continue
            idx += 1

    schedule_data = {
        "scheduleId": schedule_id,
        "start": start_state,
        "finish": finish_state,
        "items": items,
    }

    return theme_id, theme_url, theme_title, schedule_data


def _localized_value(values: dict[str, str], language: str) -> str:
    return (values.get(language, "") or "").strip()


def _build_state_for_language(
    state: ParsedState, language: str
) -> dict[str, str | None]:
    url = (state.url or "").strip()
    # Replace with mapped filename if url matches a key in DRIVE_FILENAME_MAP
    if url in DRIVE_FILENAME_MAP:
        url = DRIVE_FILENAME_MAP[url]
    return {
        "title": _localized_value(state.title, language),
        "body1": _localized_value(state.body1, language),
        "body2": _localized_value(state.body2, language),
        "url": url or None,
    }


def _convert_item_for_language(item: dict[str, Any], language: str) -> dict[str, Any]:
    converted: dict[str, Any] = {
        "itemId": item.get("itemId", ""),
        "kind": item.get("kind", "media"),
        "itemType": item.get("itemType", "image"),
        "typeId": item.get("typeId"),
        "isRecording": bool(item.get("isRecording", False)),
    }

    for state_name in ("start", "recording", "finish"):
        state = item.get(state_name)
        if isinstance(state, ParsedState):
            converted[state_name] = _build_state_for_language(state, language)

    return converted


def _build_theme_payload(
    theme_url: str,
    theme_title: dict[str, str],
    schedule_data: dict[str, Any],
    source_language: str,
) -> dict[str, Any]:
    start_state: ParsedState = schedule_data["start"]
    finish_state: ParsedState = schedule_data["finish"]
    items: list[dict[str, Any]] = schedule_data["items"]

    payload = {
        "mediaState": {
            "title": _localized_value(theme_title, source_language),
            "body1": "",
            "body2": "",
            "url": (theme_url or "").strip() or None,
        },
        "schedule": {
            "scheduleId": schedule_data.get("scheduleId") or None,
            "start": _build_state_for_language(start_state, source_language),
            "finish": _build_state_for_language(finish_state, source_language),
            "items": [
                _convert_item_for_language(item, source_language) for item in items
            ],
        },
    }
    return payload


def convert_workbook(
    input_xlsx: Path, output_root: Path, overwrite: bool = False
) -> list[Path]:
    wb = load_workbook(input_xlsx, data_only=True)
    written: list[Path] = []

    for ws in wb.worksheets:
        theme_id, theme_url, theme_title, schedule_data = _parse_sheet(ws)

        for source_lang, target_lang in LANGUAGE_EXPORT_MAP.items():
            if source_lang in IGNORED_LANGUAGES:
                continue

            payload = _build_theme_payload(
                theme_url=theme_url,
                theme_title=theme_title,
                schedule_data=schedule_data,
                source_language=source_lang,
            )

            out_dir = output_root / theme_id
            out_path = out_dir / f"{target_lang}.json"

            if out_path.exists() and not overwrite:
                raise FileExistsError(
                    f"Output file exists: {out_path} (use --overwrite to replace)"
                )

            out_dir.mkdir(parents=True, exist_ok=True)
            out_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            written.append(out_path)

    return written


def main() -> int:
    args = parse_args()

    if not args.input_xlsx.exists():
        print(f"Error: input workbook not found: {args.input_xlsx}")
        return 1

    try:
        written = convert_workbook(
            input_xlsx=args.input_xlsx,
            output_root=args.output_root,
            overwrite=args.overwrite,
        )
    except Exception as exc:
        print(f"Error: {exc}")
        return 1

    print(f"Wrote {len(written)} files")
    for path in written:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
